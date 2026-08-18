"""Generación de informes Excel con openpyxl."""
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter


class ExcelReportGenerator:
    """Generador de informes Excel profesionales."""

    COLORS = {
        "primary": "2980B9",
        "primary_dark": "1A5276",
        "accent": "85C1E9",
        "bg_light": "F0F6FC",
        "bg_table": "EAF2F8",
        "text_dark": "2C3E50",
        "success": "27AE60",
        "warning": "F39C12",
        "danger": "E74C3C",
        "white": "FFFFFF",
        "border": "D5DBDB",
    }

    def __init__(self, output_path, project_name="Instalación Eléctrica"):
        self.output_path = output_path
        self.project_name = project_name
        self.wb = Workbook()
        self._setup_styles()

    def _setup_styles(self):
        """Configurar estilos predefinidos."""
        self.header_font = Font(name='Calibri', size=12, bold=True, color=self.COLORS["white"])
        self.header_fill = PatternFill(start_color=self.COLORS["primary"], end_color=self.COLORS["primary"], fill_type='solid')
        self.title_font = Font(name='Calibri', size=16, bold=True, color=self.COLORS["primary_dark"])
        self.subtitle_font = Font(name='Calibri', size=12, bold=True, color=self.COLORS["primary"])
        self.normal_font = Font(name='Calibri', size=11, color=self.COLORS["text_dark"])
        self.result_font = Font(name='Calibri', size=11, bold=True, color=self.COLORS["primary_dark"])
        self.border = Border(
            left=Side(style='thin', color=self.COLORS["border"]),
            right=Side(style='thin', color=self.COLORS["border"]),
            top=Side(style='thin', color=self.COLORS["border"]),
            bottom=Side(style='thin', color=self.COLORS["border"]),
        )
        self.center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        self.left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
        self.row_fill_even = PatternFill(start_color=self.COLORS["bg_table"], end_color=self.COLORS["bg_table"], fill_type='solid')
        self.row_fill_odd = PatternFill(start_color=self.COLORS["white"], end_color=self.COLORS["white"], fill_type='solid')
        self.success_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type='solid')
        self.danger_fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type='solid')

    def add_title_sheet(self):
        """Añadir hoja de portada."""
        ws = self.wb.active
        ws.title = "Portada"
        ws.merge_cells('B2:H2')
        ws['B2'] = "ELECTRICITRON"
        ws['B2'].font = Font(name='Calibri', size=28, bold=True, color=self.COLORS["primary_dark"])
        ws['B2'].alignment = Alignment(horizontal='center')
        ws.merge_cells('B4:H4')
        ws['B4'] = self.project_name
        ws['B4'].font = self.title_font
        ws['B4'].alignment = Alignment(horizontal='center')
        ws.merge_cells('B6:H6')
        ws['B6'] = f"Fecha: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws['B6'].font = self.normal_font
        ws['B6'].alignment = Alignment(horizontal='center')
        ws.merge_cells('B8:H8')
        ws['B8'] = "Software de cálculos eléctricos y telecomunicaciones"
        ws['B8'].font = Font(name='Calibri', size=11, italic=True, color=self.COLORS["text_dark"])
        ws['B8'].alignment = Alignment(horizontal='center')
        for col in range(2, 9):
            ws.column_dimensions[get_column_letter(col)].width = 16

    def add_data_sheet(self, sheet_name, headers, data, title=None):
        """Añadir hoja con datos."""
        ws = self.wb.create_sheet(title=sheet_name[:31])
        row = 1
        if title:
            ws.merge_cells(f'A{row}:{get_column_letter(len(headers))}{row}')
            ws[f'A{row}'] = title
            ws[f'A{row}'].font = self.title_font
            ws[f'A{row}'].alignment = Alignment(horizontal='center')
            row += 2

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_align
            cell.border = self.border
        row += 1

        for data_row_idx, data_row in enumerate(data):
            fill = self.row_fill_even if data_row_idx % 2 == 0 else self.row_fill_odd
            for col_idx, value in enumerate(data_row, 1):
                cell = ws.cell(row=row, column=col_idx, value=value)
                cell.font = self.normal_font
                cell.fill = fill
                cell.border = self.border
                cell.alignment = self.center_align if col_idx > 1 else self.left_align
            row += 1

        for col_idx in range(1, len(headers) + 1):
            max_length = max(
                len(str(ws.cell(row=r, column=col_idx).value or ""))
                for r in range(1, row)
            )
            ws.column_dimensions[get_column_letter(col_idx)].width = max(max_length + 4, 15)

        ws.auto_filter.ref = f"A{row - len(data) - 1}:{get_column_letter(len(headers))}{row - 1}"

    def add_results_sheet(self, sheet_name, results_dict, title="Resultados"):
        """Añadir hoja de resultados key-value."""
        ws = self.wb.create_sheet(title=sheet_name[:31])
        row = 1
        ws.merge_cells(f'A{row}:B{row}')
        ws[f'A{row}'] = title
        ws[f'A{row}'].font = self.title_font
        ws[f'A{row}'].alignment = Alignment(horizontal='center')
        row += 2

        headers = ["Parámetro", "Valor"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_align
            cell.border = self.border
        row += 1

        for key, value in results_dict.items():
            clean_key = key.replace("_", " ").title()
            fill = self.row_fill_even if (row % 2 == 0) else self.row_fill_odd
            cell_k = ws.cell(row=row, column=1, value=clean_key)
            cell_k.font = self.normal_font
            cell_k.fill = fill
            cell_k.border = self.border
            cell_k.alignment = self.left_align

            cell_v = ws.cell(row=row, column=2, value=str(value))
            cell_v.font = self.result_font
            cell_v.fill = fill
            cell_v.border = self.border
            cell_v.alignment = self.center_align

            if isinstance(value, bool):
                cell_v.fill = self.success_fill if value else self.danger_fill

            row += 1

        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 25

    def save(self):
        """Guardar el archivo Excel."""
        self.wb.save(self.output_path)
        return self.output_path


class ExcelReportManager:
    """Gestión de datos en informes Excel."""

    def __init__(self):
        self.records = []
        self.current_id = 0

    def add_record(self, category, title, params, results, notes=None):
        """Añadir registro."""
        self.current_id += 1
        record = {
            "id": self.current_id,
            "category": category,
            "title": title,
            "params": params,
            "results": results,
            "notes": notes or [],
            "timestamp": datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        }
        self.records.append(record)
        return record

    def modify_record(self, record_id, **kwargs):
        """Modificar registro."""
        for record in self.records:
            if record["id"] == record_id:
                for key, value in kwargs.items():
                    if key in record:
                        record[key] = value
                return record
        return None

    def delete_record(self, record_id):
        """Eliminar registro."""
        self.records = [r for r in self.records if r["id"] != record_id]

    def get_records_by_category(self, category):
        """Obtener registros por categoría."""
        return [r for r in self.records if r["category"] == category]

    def get_all_records(self):
        """Obtener todos los registros."""
        return self.records

    def clear_all(self):
        """Limpiar registros."""
        self.records.clear()

    def generate_excel(self, output_path, project_name="Instalación Eléctrica"):
        """Generar archivo Excel con todos los registros."""
        gen = ExcelReportGenerator(output_path, project_name)
        gen.add_title_sheet()

        categories = {}
        for record in self.records:
            cat = record["category"]
            if cat not in categories:
                categories[cat] = []
            categories[cat].append(record)

        category_names = {
            "basico": "Cálculos Básicos",
            "cable": "Cables",
            "proteccion": "Protecciones",
            "instalacion": "Instalaciones",
            "telecom": "Telecomunicaciones",
            "distancia": "Distancias",
        }

        for cat, records in categories.items():
            cat_name = category_names.get(cat, cat.title())
            headers = ["ID", "Título", "Parámetros", "Resultados", "Notas", "Fecha"]
            data = []
            for rec in records:
                params_str = "; ".join(f"{k}: {v}" for k, v in rec["params"].items()) if rec["params"] else ""
                results_str = "; ".join(f"{k}: {v}" for k, v in rec["results"].items()) if rec["results"] else ""
                notes_str = "; ".join(rec["notes"]) if rec["notes"] else ""
                data.append([
                    rec["id"], rec["title"], params_str, results_str, notes_str, rec["timestamp"]
                ])
            gen.add_data_sheet(cat_name, headers, data, title=cat_name)

        if self.records:
            ws_summary = gen.wb.create_sheet(title="Resumen", index=1)
            ws_summary.merge_cells('A1:C1')
            ws_summary['A1'] = "Resumen de Cálculos"
            ws_summary['A1'].font = gen.title_font
            ws_summary['A1'].alignment = Alignment(horizontal='center')

            summary_headers = ["Categoría", "Nº de Cálculos", "Fecha"]
            for col_idx, h in enumerate(summary_headers, 1):
                cell = ws_summary.cell(row=3, column=col_idx, value=h)
                cell.font = gen.header_font
                cell.fill = gen.header_fill
                cell.alignment = gen.center_align
                cell.border = gen.border

            row = 4
            for cat, records in categories.items():
                ws_summary.cell(row=row, column=1, value=category_names.get(cat, cat)).border = gen.border
                ws_summary.cell(row=row, column=2, value=len(records)).border = gen.border
                ws_summary.cell(row=row, column=3, value=datetime.now().strftime('%d/%m/%Y')).border = gen.border
                row += 1

            ws_summary.column_dimensions['A'].width = 30
            ws_summary.column_dimensions['B'].width = 20
            ws_summary.column_dimensions['C'].width = 18

        gen.save()
        return output_path
